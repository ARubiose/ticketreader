"""Excel adapter for tickets."""
import pathlib
import logging
import functools
from typing import Optional, List, Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.cell.cell import Cell

from ticketreader.schemas import Ticket, UnitProduct, BulkProduct

logging = logging.getLogger(__name__)

DEFAULT_TABLE_STYLE = TableStyleInfo(name="DefaultTableStyle", showFirstColumn=False,
                                     showLastColumn=False, showRowStripes=True, showColumnStripes=True)
UNIT_PRODUCT_TABLE_HEADER = ["Supermercado", "Fecha de compra",
                             "ID de factura", "Cantidad", "Producto", "Precio"]
TABLES_START_CELL = "C3"


class ExcelHandler:
    """Excel adapter for tickets."""

    # Sheet names
    TICKETS_SHEET: str = "Tickets"
    UNIT_PRODUCT_SHEET: str = "Productos unitarios"
    BULK_PRODUCT_SHEET: str = "Productos a granel"

    # Range
    TABLES_START_ROW = 4
    TABLES_START_COLUMN = 2
    TABLE_TOTAL_COLUMN = 12
    UNIT_PRODUCT_LAST_COLUMN = 11
    BULK_PRODUCT_LAST_COLUMN = 12
    TICKET_IVA_LAST_COLUMN = 10
    TICKET_TOTAL_LAST_COLUMN = 16

    def __init__(self, file_name: pathlib.Path):
        """Initialize Excel adapter."""
        self._file_name = file_name
        self.workbook = file_name

    @property
    def workbook(self) -> Workbook:
        """Workbook instance."""
        return self._workbook

    @workbook.setter
    def workbook(self, file_name: pathlib.Path) -> None:
        """Workbook instance."""
        try:
            self._workbook = load_workbook(filename=file_name)
        except FileNotFoundError:
            self._workbook = Workbook()

    @property
    def unit_product_sheet(self) -> Worksheet:
        """Unit product sheet."""
        return self.get_sheet(self.UNIT_PRODUCT_SHEET)

    @property
    def bulk_product_sheet(self) -> Worksheet:
        """Bulk product sheet."""
        return self.get_sheet(self.BULK_PRODUCT_SHEET)

    @property
    def tickets_sheet(self) -> Worksheet:
        """Tickets sheet."""
        return self.get_sheet(self.TICKETS_SHEET)

    def get_sheet(self, sheet_name: str) -> Worksheet:
        """Get sheet by name or create it if it doesn't exist"""
        try:
            return self.workbook[sheet_name]
        except KeyError:
            return self.workbook.create_sheet(title=sheet_name)

    def save_ticket(self, ticket: Ticket):
        """Save tickets to an Excel file."""

        self._save_unit_products(ticket)
        self._save_bulk_products(ticket)
        self._save_total(ticket)

        # Save workbook
        self.save_workbook()

    def _save_unit_products(self, ticket: Ticket):
        """Save unit products to an Excel file."""

        for product_row in self.unit_product_list_generator(ticket):
            self.unit_product_sheet.insert_rows(self.TABLES_START_ROW)
            new_row = self.unit_product_sheet[self.TABLES_START_ROW]
            new_table_row = new_row[self.TABLES_START_COLUMN:self.UNIT_PRODUCT_LAST_COLUMN]
            for cell, value in zip(new_table_row, product_row):
                cell.value = value

    def unit_product_list_generator(self, ticket: Ticket) -> Iterator[List[str | int | float]]:
        """Build unit product list."""
        unit_products = [
            p for p in ticket.products if isinstance(p, UnitProduct)]
        for product in unit_products:
            yield [
                ticket.supermarket.id.value,
                f'{ticket.supermarket.address.street}, {ticket.supermarket.address.postal_code}, {ticket.supermarket.address.city}',
                ticket.purchase_datetime.strftime("%d/%m/%Y %H:%M:%S"),
                ticket.invoice_id,
                product.quantity,
                product.name,
                product.price_per_item,
                product.total
            ]

        return UNIT_PRODUCT_TABLE_HEADER

    def _save_bulk_products(self, ticket: Ticket):
        """Save bulk products to an Excel file."""

        for product_row in self.bulk_product_list_generator(ticket):
            self.bulk_product_sheet.insert_rows(self.TABLES_START_ROW)
            new_row = self.bulk_product_sheet[self.TABLES_START_ROW]
            new_table_row = new_row[self.TABLES_START_COLUMN:self.UNIT_PRODUCT_LAST_COLUMN]
            for cell, value in zip(new_table_row, product_row):
                cell.value = value

    def bulk_product_list_generator(self, ticket: Ticket) -> Iterator[List[str | int | float]]:
        """Build bulk product list."""
        bulk_products = [
            p for p in ticket.products if isinstance(p, BulkProduct)]
        for product in bulk_products:
            yield [
                ticket.supermarket.id.value,
                f'{ticket.supermarket.address.street}, {ticket.supermarket.address.postal_code}, {ticket.supermarket.address.city}',
                ticket.purchase_datetime.strftime("%d/%m/%Y %H:%M:%S"),
                ticket.invoice_id,
                product.quantity,
                product.name,
                product.unit_of_measure,
                product.price_per_unit,
                product.total
            ]

        return UNIT_PRODUCT_TABLE_HEADER

    def _save_total(self, ticket: Ticket):
        """Save ticket total to an Excel file."""

        self.tickets_sheet.insert_rows(self.TABLES_START_ROW)
        new_row = self.tickets_sheet[self.TABLES_START_ROW]
        total_row = self._get_total_row(ticket)
        new_table_row = new_row[self.TABLES_START_COLUMN:self.TICKET_TOTAL_LAST_COLUMN]
        for cell, value in zip(new_table_row, total_row):
            cell.value = value

    def _get_total_row(self, ticket: Ticket) -> List[str | float]:
        """Get total row."""
        total_vat = [
            iva_item.fee for iva_item in ticket.iva if iva_item.type == "TOTAL"]
        return [
            ticket.supermarket.id.value,
            f'{ticket.supermarket.address.street}, {ticket.supermarket.address.postal_code}, {ticket.supermarket.address.city}',
            ticket.purchase_datetime.strftime("%d/%m/%Y %H:%M:%S"),
            ticket.invoice_id,
            total_vat[0] if total_vat else 0,
            ticket.total,
        ]

    def save_workbook(self):
        """Save workbook to file."""
        self.workbook.save(self._file_name)
